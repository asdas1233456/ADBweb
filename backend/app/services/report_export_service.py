"""
测试报告导出服务
支持导出为PDF、Excel、Word等多种格式
"""
from sqlmodel import Session, select
from app.models.task_log import TaskLog
from app.models import Device, Script
from typing import List, Dict, Optional
import logging
from datetime import datetime
from io import BytesIO
import json

logger = logging.getLogger(__name__)


class ReportExportService:
    """测试报告导出服务"""
    
    def __init__(self, session: Session):
        """
        初始化报告导出服务
        
        Args:
            session: 数据库会话
        """
        self.session = session
    
    def export_to_excel(
        self, 
        task_log_ids: List[int],
        include_details: bool = True
    ) -> BytesIO:
        """
        导出为Excel格式
        
        Args:
            task_log_ids: 任务日志ID列表
            include_details: 是否包含详细信息
            
        Returns:
            Excel文件的BytesIO对象
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl未安装，请运行: pip install openpyxl")
            raise ImportError("需要安装openpyxl库")
        
        # 获取任务日志
        task_logs = self._get_task_logs(task_log_ids)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试报告"
        
        # 设置标题样式
        title_font = Font(bold=True, size=12)
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入标题行
        headers = [
            "任务ID", "任务名称", "脚本名称", "设备名称", 
            "状态", "开始时间", "结束时间", "执行时长(秒)", "错误信息"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
        
        # 写入数据
        for row_num, task_log in enumerate(task_logs, 2):
            # 获取关联数据
            script = self.session.get(Script, task_log.script_id) if task_log.script_id else None
            device = self.session.get(Device, task_log.device_id) if task_log.device_id else None
            
            # 计算执行时长
            duration = None
            if task_log.start_time and task_log.end_time:
                duration = (task_log.end_time - task_log.start_time).total_seconds()
            
            # 写入行数据
            row_data = [
                task_log.id,
                task_log.task_name or "",
                script.name if script else "",
                device.model if device else "",
                self._get_status_text(task_log.status),
                task_log.start_time.strftime("%Y-%m-%d %H:%M:%S") if task_log.start_time else "",
                task_log.end_time.strftime("%Y-%m-%d %H:%M:%S") if task_log.end_time else "",
                round(duration, 2) if duration else "",
                task_log.error_message or ""
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="center")
                
                # 根据状态设置颜色
                if col_num == 5:  # 状态列
                    if task_log.status == 'success':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif task_log.status == 'failed':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 自动调整列宽
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # 添加统计信息工作表
        if include_details:
            self._add_statistics_sheet(wb, task_logs)
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel报告生成成功，包含 {len(task_logs)} 条记录")
        return output
    
    def export_to_pdf(
        self, 
        task_log_ids: List[int],
        include_charts: bool = True
    ) -> BytesIO:
        """
        导出为PDF格式
        
        Args:
            task_log_ids: 任务日志ID列表
            include_charts: 是否包含图表
            
        Returns:
            PDF文件的BytesIO对象
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            logger.error("reportlab未安装，请运行: pip install reportlab")
            raise ImportError("需要安装reportlab库")
        
        # 获取任务日志
        task_logs = self._get_task_logs(task_log_ids)
        
        # 创建PDF文档
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1890ff'),
            spaceAfter=30,
            alignment=1  # 居中
        )
        
        # 标题
        title = Paragraph("测试执行报告", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2*inch))
        
        # 报告信息
        report_info = [
            ["报告生成时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["测试任务数量:", str(len(task_logs))],
            ["成功任务:", str(sum(1 for t in task_logs if t.status == 'success'))],
            ["失败任务:", str(sum(1 for t in task_logs if t.status == 'failed'))],
        ]
        
        info_table = Table(report_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 任务详情表格
        table_data = [["任务ID", "任务名称", "状态", "开始时间", "执行时长"]]
        
        for task_log in task_logs:
            duration = ""
            if task_log.start_time and task_log.end_time:
                duration = f"{(task_log.end_time - task_log.start_time).total_seconds():.2f}s"
            
            table_data.append([
                str(task_log.id),
                task_log.task_name or "",
                self._get_status_text(task_log.status),
                task_log.start_time.strftime("%Y-%m-%d %H:%M:%S") if task_log.start_time else "",
                duration
            ])
        
        # 创建表格
        task_table = Table(table_data, colWidths=[0.8*inch, 2*inch, 1*inch, 1.8*inch, 1*inch])
        task_table.setStyle(TableStyle([
            # 标题行样式
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # 数据行样式
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        
        story.append(task_table)
        
        # 构建PDF
        doc.build(story)
        output.seek(0)
        
        logger.info(f"PDF报告生成成功，包含 {len(task_logs)} 条记录")
        return output
    
    def export_to_json(
        self, 
        task_log_ids: List[int],
        include_logs: bool = False
    ) -> str:
        """
        导出为JSON格式
        
        Args:
            task_log_ids: 任务日志ID列表
            include_logs: 是否包含详细日志
            
        Returns:
            JSON字符串
        """
        task_logs = self._get_task_logs(task_log_ids)
        
        report_data = {
            'generated_at': datetime.now().isoformat(),
            'total_tasks': len(task_logs),
            'summary': self._generate_summary(task_logs),
            'tasks': []
        }
        
        for task_log in task_logs:
            script = self.session.get(Script, task_log.script_id) if task_log.script_id else None
            device = self.session.get(Device, task_log.device_id) if task_log.device_id else None
            
            task_data = {
                'id': task_log.id,
                'task_name': task_log.task_name,
                'script_name': script.name if script else None,
                'device_name': device.model if device else None,
                'status': task_log.status,
                'start_time': task_log.start_time.isoformat() if task_log.start_time else None,
                'end_time': task_log.end_time.isoformat() if task_log.end_time else None,
                'error_message': task_log.error_message,
            }
            
            if include_logs:
                task_data['log_content'] = task_log.log_content
            
            report_data['tasks'].append(task_data)
        
        logger.info(f"JSON报告生成成功，包含 {len(task_logs)} 条记录")
        return json.dumps(report_data, ensure_ascii=False, indent=2)
    
    def export_to_html(
        self, 
        task_log_ids: List[int]
    ) -> str:
        """
        导出为HTML格式
        
        Args:
            task_log_ids: 任务日志ID列表
            
        Returns:
            HTML字符串
        """
        task_logs = self._get_task_logs(task_log_ids)
        summary = self._generate_summary(task_logs)
        
        html = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>测试执行报告</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #1890ff;
            text-align: center;
            margin-bottom: 30px;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 20px;
            margin-bottom: 30px;
        }}
        .summary-card {{
            padding: 20px;
            border-radius: 4px;
            text-align: center;
        }}
        .summary-card.total {{
            background-color: #e6f7ff;
            border: 1px solid #91d5ff;
        }}
        .summary-card.success {{
            background-color: #f6ffed;
            border: 1px solid #b7eb8f;
        }}
        .summary-card.failed {{
            background-color: #fff1f0;
            border: 1px solid #ffa39e;
        }}
        .summary-card.rate {{
            background-color: #fff7e6;
            border: 1px solid #ffd591;
        }}
        .summary-card h3 {{
            margin: 0 0 10px 0;
            font-size: 14px;
            color: #666;
        }}
        .summary-card .value {{
            font-size: 32px;
            font-weight: bold;
            margin: 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #4472C4;
            color: white;
            font-weight: bold;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .status {{
            padding: 4px 12px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: bold;
        }}
        .status.success {{
            background-color: #52c41a;
            color: white;
        }}
        .status.failed {{
            background-color: #ff4d4f;
            color: white;
        }}
        .status.running {{
            background-color: #1890ff;
            color: white;
        }}
        .footer {{
            margin-top: 30px;
            text-align: center;
            color: #999;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 测试执行报告</h1>
        
        <div class="summary">
            <div class="summary-card total">
                <h3>总任务数</h3>
                <p class="value">{summary['total']}</p>
            </div>
            <div class="summary-card success">
                <h3>成功任务</h3>
                <p class="value">{summary['success']}</p>
            </div>
            <div class="summary-card failed">
                <h3>失败任务</h3>
                <p class="value">{summary['failed']}</p>
            </div>
            <div class="summary-card rate">
                <h3>成功率</h3>
                <p class="value">{summary['success_rate']:.1f}%</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>任务ID</th>
                    <th>任务名称</th>
                    <th>脚本名称</th>
                    <th>设备名称</th>
                    <th>状态</th>
                    <th>开始时间</th>
                    <th>执行时长</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for task_log in task_logs:
            script = self.session.get(Script, task_log.script_id) if task_log.script_id else None
            device = self.session.get(Device, task_log.device_id) if task_log.device_id else None
            
            duration = ""
            if task_log.start_time and task_log.end_time:
                duration = f"{(task_log.end_time - task_log.start_time).total_seconds():.2f}s"
            
            status_class = task_log.status
            status_text = self._get_status_text(task_log.status)
            
            html += f"""
                <tr>
                    <td>{task_log.id}</td>
                    <td>{task_log.task_name or ''}</td>
                    <td>{script.name if script else ''}</td>
                    <td>{device.model if device else ''}</td>
                    <td><span class="status {status_class}">{status_text}</span></td>
                    <td>{task_log.start_time.strftime("%Y-%m-%d %H:%M:%S") if task_log.start_time else ''}</td>
                    <td>{duration}</td>
                </tr>
"""
        
        html += f"""
            </tbody>
        </table>
        
        <div class="footer">
            <p>报告生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
            <p>ADBweb - Android 自动化测试平台</p>
        </div>
    </div>
</body>
</html>
"""
        
        logger.info(f"HTML报告生成成功，包含 {len(task_logs)} 条记录")
        return html
    
    def _get_task_logs(self, task_log_ids: List[int]) -> List[TaskLog]:
        """获取任务日志列表"""
        statement = select(TaskLog).where(TaskLog.id.in_(task_log_ids))
        return self.session.exec(statement).all()
    
    def _get_status_text(self, status: str) -> str:
        """获取状态文本"""
        status_map = {
            'success': '成功',
            'failed': '失败',
            'running': '运行中',
            'pending': '等待中'
        }
        return status_map.get(status, status)
    
    def _generate_summary(self, task_logs: List[TaskLog]) -> Dict:
        """生成统计摘要"""
        total = len(task_logs)
        success = sum(1 for t in task_logs if t.status == 'success')
        failed = sum(1 for t in task_logs if t.status == 'failed')
        success_rate = (success / total * 100) if total > 0 else 0
        
        return {
            'total': total,
            'success': success,
            'failed': failed,
            'success_rate': success_rate
        }
    
    def _add_statistics_sheet(self, workbook, task_logs: List[TaskLog]):
        """添加统计信息工作表"""
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.chart import PieChart, Reference
        
        ws = workbook.create_sheet("统计信息")
        
        # 标题
        ws['A1'] = "测试执行统计"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # 统计数据
        summary = self._generate_summary(task_logs)
        stats_data = [
            ["总任务数", summary['total']],
            ["成功任务", summary['success']],
            ["失败任务", summary['failed']],
            ["成功率", f"{summary['success_rate']:.2f}%"]
        ]
        
        for row_num, (label, value) in enumerate(stats_data, 3):
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
